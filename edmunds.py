from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.touch_actions import TouchActions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from fake_useragent import UserAgent
import time
import openpyxl






def configure_driver():
    # Add additional Options to the webdriver
    chrome_options = Options()
    ua = UserAgent()
    userAgent = ua.random                                     #THIS IS FAKE AGENT IT WILL GIVE YOU NEW AGENT EVERYTIME
    print(userAgent)
   # add the argument and make the browser Headless.
   # chrome_options.add_argument("--headless")                    if you don't want to see the display on chrome just uncomment this
    chrome_options.add_argument(f'user-agent={userAgent}')     #useragent added
    chrome_options.add_argument("--log-level=3")               #removes error/warning/info messages displayed on the console
    chrome_options.add_argument("--disable-notifications")     #disable notifications
    chrome_options.add_argument("--disable-infobars")         #disable infobars ""Chrome is being controlled by automated test software"  Although is isn't supported by Chrome anymore
    chrome_options.add_argument("start-maximized")            #will maximize chrome screen
    chrome_options.add_argument('--disable-gpu')             #disable gpu (not load pictures fully)
    chrome_options.add_argument("--disable-extensions")       #will disable developer mode extensions
    #chrome_options.add_argument('--proxy-server=%s' % PROXY)
    prefs = {"profile.managed_default_content_settings.images": 2}
    chrome_options.add_experimental_option("prefs", prefs)             #we have disabled pictures (so no time is wasted in loading them)
    driver = webdriver.Chrome(ChromeDriverManager().install(), options = chrome_options)   #you don't have to download chromedriver it will be downloaded by itself and will be saved in cache
    return driver

def RunScrapper(driver):

    start_time = time.time()


    #workbook created
    wb = openpyxl.Workbook()
    # add_sheet is used to create sheet.
    sheet1 = wb.active
    print(" WORKSHEET CREATED SUCCESSFULLY!")
    # INITIALIZING THE COLOUMN NAMES NOW
    c1 = sheet1.cell(row = 1, column = 1)
    c1.value= "Name"
    c2 = sheet1.cell(row=1, column=2)
    c2.value = "Price"
    c3 = sheet1.cell(row=1, column=3)
    c3.value = "VIN"
    c4 = sheet1.cell(row=1, column=4)
    c4.value = "Vechicle Summary"
    c5 = sheet1.cell(row=1, column=5)
    c5.value = "Top Features & Specs"
    wb.save("edmunds.xlsx")
    #setting row number to 2
    mi=2


    #getting input from user
    print("                   ")
    print("NOTE: number of pages option is given so that it doesn't take time for you to check the code. You can take it as >600 if you want to check all pages")
    print("                   ")
    print("                   ")
    zip = input("Enter Zip Code: ")
    numpg=int(input("Enter the number of pages you want to parse: "))


    #going to website
    driver.get("https://www.edmunds.com/cars-for-sale-by-owner/")
    #Waiting for input field of zipcode, backspacing the exisiting number and putting user provided zipcode
    WebDriverWait(driver,20).until(expected_conditions.visibility_of_element_located((By.XPATH,'//input[@name="zip"]')))
    zipcode=driver.find_element_by_xpath('//input[@name="zip"]')
    for i in range(8):
        zipcode.send_keys(Keys.BACKSPACE)
    zipcode.send_keys(zip)
    zipcode.send_keys(Keys.ENTER)


    #this section parsing number of pages said by user and takes links of each entry in a list

    links = []
    for pg in range(numpg):
        WebDriverWait(driver,20).until(expected_conditions.invisibility_of_element_located((By.XPATH,'//div[@class="loading-overlay"]')))
        WebDriverWait(driver, 40).until(expected_conditions.visibility_of_element_located(
            (By.XPATH, '//ul[@data-tracking-parent="edm-entry-vehicle-card-list"]')))
        mainbl = driver.find_element_by_xpath('//ul[@data-tracking-parent="edm-entry-vehicle-card-list"]')
        container = mainbl.find_elements_by_xpath('.//li[@class="d-flex mb-0_75 mb-md-1_5 col-12 col-md-6"]')
        print("Links On This Page: ",len(container))
        for contain in container:
            linka = contain.find_element_by_tag_name('a').get_attribute('href')
            links.append(linka)
        print("Total Links : ", len(links))
        try:
            #incase next button is there it will click or else get out of for loop
            driver.find_element_by_xpath('//a[@aria-label="Pagination right"]').click()
        except Exception:
            break

    entryno =1

    #this section scrapes each entry 1 by 1
    #I have put try and except at each field so incase if it isn't available on page it can be ignored
    #a whole try except is also provided so incase if link is broken we move onto next link
    for l in links:

        print("Link: ",l)
        print("Entry No: ",entryno)
        entryno=entryno+1

        try:
            driver.get(l)

            #incase we don't see overview and wait exceeds 30 seconds link will be considered broken

            WebDriverWait(driver, 30).until(expected_conditions.visibility_of_element_located((By.NAME, 'overview')))

            overview = driver.find_element_by_name('overview')

            #Name section

            NAME = overview.find_element_by_tag_name('h1').text

            #Price section

            try:
                PRICE = driver.find_element_by_xpath('//span[@data-test="vdp-price-row"]').text
            except Exception:
                PRICE = ''
                pass

            #VIN section

            try:
                VIN = driver.find_element_by_class_name('mr-1').text
                VIN = VIN.replace("VIN: ", "")
            except Exception:
                VIN = ''
                pass


            # vehicle summary section

            try:
                vehiclesummary = driver.find_element_by_xpath(
                    '//section[@data-tracking-parent="edm-entry-vehicle-summary"]')
                vehiclecontainer = vehiclesummary.find_elements_by_xpath('.//div[@class="m-0 mb-1 row"]')
                vechiclesum = []

                for vcontain in vehiclecontainer:
                    vechiclesum.append(vcontain.text)
            except Exception:
                vechiclesum = ''
                pass

            #top specs section

            try:
                topspecsmain = driver.find_element_by_xpath(
                    '//section[@data-tracking-parent="edm-entry-features-specs"]')
                topcontainer = topspecsmain.find_elements_by_xpath('.//div[@class="mb-1 col-12 col-md-6"]')
                topspec = []
                keys = []
                values = []
                dict = {}
                for t in topcontainer:
                    heading = t.find_element_by_tag_name('div').text
                    keys.append(heading)
                    try:
                        t.find_element_by_tag_name('button').click()
                    except Exception:
                        pass
                    tcontainer = t.find_elements_by_class_name('mb-0_5')
                    p = 1
                    val = ''
                    for tc in tcontainer:
                        if p == 1:
                            val = tc.text
                        else:
                            val = val + ',' + tc.text
                        p = p + 1
                    values.append(val)
                for k in range(len(keys)):
                    torep=keys[k]+','
                    valme=values[k].replace(torep,"")
                    try:
                        valme=valme.replace(",See less","")
                    except Exception:
                        pass
                    dict[keys[k]] = valme
            except Exception:
                dict = {}
                pass

            #print section

            print("Name: ", NAME)
            print("Price: ", PRICE)
            print("VIN: ", VIN)
            print("Vehicle Summary: ", vechiclesum)
            print("Top Specs: ",dict)

            #saving in excel section

            c1 = sheet1.cell(row=mi, column=1)
            c1.value = NAME
            c2 = sheet1.cell(row=mi, column=2)
            c2.value = PRICE
            c3 = sheet1.cell(row=mi, column=3)
            c3.value = VIN
            vechiclesum=str(vechiclesum)
            c4 = sheet1.cell(row=mi, column=4)
            c4.value = vechiclesum
            dict=str(dict)
            c5 = sheet1.cell(row=mi, column=5)
            c5.value = dict
            wb.save("edmunds.xlsx")
            mi = mi+1
            print("")
            print("**************")
            print("")
        except Exception:
            print("Link Broken")
            pass

    #give time taken to execute everything
    print("time elapsed: {:.2f}s".format(time.time() - start_time))



# create the driver object.
driver= configure_driver()

#call the scrapper to run
RunScrapper(driver)

# close the driver.
driver.close()














